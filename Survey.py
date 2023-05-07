import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# These libraries are required to load the existing worksheet and insert the image in it
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO

from scipy.stats import f_oneway

# Load the dataset
df = pd.read_csv("survey_results_public.csv")

# Create a new DataFrame to store all the responses
all_responses = pd.DataFrame(columns=["Work Arrangement", "Job Satisfaction", "Age", "Gender"])

# Ask the user questions and append the responses to the DataFrame
while True:
    work_arrangement = input("What is your work arrangement (remote/hybrid/face-to-face)? ").lower() 
    if work_arrangement not in ["remote", "hybrid", "face-to-face"]:
        print("Invalid work arrangement. Please enter either remote, hybrid, or face-to-face.")
        continue

    # The input function for the job satisfaction rating is validated to ensure 
    # that the user enters a number between 1 and 10. 
    job_satisfaction = input("On a scale of 1-10, how satisfied are you with your job? ")
    try:
        job_satisfaction = int(job_satisfaction)
        if not 1 <= job_satisfaction <= 10:
            raise ValueError
    except ValueError:
        print("Invalid job satisfaction rating. Please enter a number between 1 and 10.")
        continue
    age = input("What is your age group? ")
    try:
        age = int(age)
        if age < 18:
            print("You must be 18 years or older to participate in this survey.")
            continue
    except ValueError:
        print("Invalid age. Please enter a number.")
        continue
    gender = input("What is your gender? ").lower()
    if gender not in ["male", "female"]:
        print("Invalid gender. Please enter either male or female.")
        continue
    
    # Append the responses to the DataFrame
    new_row = {"Work Arrangement": work_arrangement,
               "Job Satisfaction": job_satisfaction,
               "Age": age,
               "Gender": gender}
    df = pd.concat([df, pd.DataFrame(new_row, index=[0])], ignore_index=True)
    
    # Ask if the user wants to continue
    another_response = input("Do you want to add another response? (y/n)").lower()
    if another_response == "n":
        break

# Append the new responses to the existing responses in the Excel file
with pd.ExcelWriter('survey_results_public.xlsx') as writer:
    df.to_excel(writer, sheet_name='Data', startrow=0, startcol=0)

# Filter the dataset to include only respondents aged 18 or older and with valid work arrangement and gender values
df_filtered = df[(df["Age"] >= 18) & 
                 (df["Work Arrangement"].isin(["Remote", "Hybrid", "Face-to-face"])) & 
                 (df["Gender"].isin(["Male", "Female"]))]

# Compute the mean job satisfaction and standard error of the mean for each work arrangement
remote_satisfaction = np.mean(df_filtered[df_filtered["Work Arrangement"] == "Remote"]["Job Satisfaction"])
hybrid_satisfaction = np.mean(df_filtered[df_filtered["Work Arrangement"] == "Hybrid"]["Job Satisfaction"])
face_to_face_satisfaction = np.mean(df_filtered[df_filtered["Work Arrangement"] == "Face-to-face"]["Job Satisfaction"])
remote_se = np.std(df_filtered[df_filtered["Work Arrangement"] == "Remote"]["Job Satisfaction"]) / np.sqrt(len(df_filtered[df_filtered["Work Arrangement"] == "Remote"]))
hybrid_se = np.std(df_filtered[df_filtered["Work Arrangement"] == "Hybrid"]["Job Satisfaction"]) / np.sqrt(len(df_filtered[df_filtered["Work Arrangement"] == "Hybrid"]))
face_to_face_se = np.std(df_filtered[df_filtered["Work Arrangement"] == "Face-to-face"]["Job Satisfaction"]) / np.sqrt(len(df_filtered[df_filtered["Work Arrangement"] == "Face-to-face"]))

# Create a grouped bar chart of mean job satisfaction by work arrangement
work_arrangements = ["Remote", "Hybrid", "Face-to-face"]
satisfactions = [remote_satisfaction, hybrid_satisfaction, face_to_face_satisfaction]
ses = [remote_se, hybrid_se, face_to_face_se]

fig, ax = plt.subplots()
ax.bar(work_arrangements, satisfactions, yerr=ses)
ax.set_xlabel("Work Arrangement")
ax.set_ylabel("Mean Job Satisfaction")
ax.set_title("Job Satisfaction by Work Arrangement")

# Save the chart in the Excel file
with pd.ExcelWriter('survey_results_public.xlsx', engine='openpyxl', mode='a') as writer:

    # Load the existing worksheet
    writer.book = load_workbook('survey_results_public.xlsx')
    sheet_name = 'Data'
    if sheet_name in writer.book.sheetnames:
        worksheet = writer.book[sheet_name]
    else:
        worksheet = writer.book.create_sheet(sheet_name)

    # Add the chart to the worksheet
    img = plt.savefig('job_satisfaction_by_work_arrangement.png')
    img_file = open('job_satisfaction_by_work_arrangement.png', 'rb').read()
    img_data = BytesIO(img_file)
    img = Image(img_data)
    worksheet.add_image(img, 'H1')

# Create separate dataframes for each work arrangement
remote_df = df_filtered[df_filtered["Work Arrangement"] == "Remote"]
hybrid_df = df_filtered[df_filtered["Work Arrangement"] == "Hybrid"]
face_to_face_df = df_filtered[df_filtered["Work Arrangement"] == "Face-to-face"]

# Conduct one-way ANOVA test
f_stat, p_val = f_oneway(df_filtered[df_filtered["Work Arrangement"] == "Remote"]["Job Satisfaction"],
                          df_filtered[df_filtered["Work Arrangement"] == "Hybrid"]["Job Satisfaction"],
                          df_filtered[df_filtered["Work Arrangement"] == "Face-to-face"]["Job Satisfaction"])

print("F-statistic: {:.2f}".format(f_stat))
print("p-value: {:.4f}".format(p_val))

if p_val < 0.05:
    print("Reject the null hypothesis. There is a significant difference in mean job satisfaction scores among employees in different work arrangements.")
else:
    print("Fail to reject the null hypothesis. There is no significant difference in mean job satisfaction scores among employees in different work arrangements.")
